from openpyxl import load_workbook


# transferring data from excel to array
def gendata(startrow, endrow,col,filepath,sheetname):
    # Load the work book
    wb = load_workbook(filepath, read_only=False, data_only=True)
    # find out the sheets
    ws = wb.sheetnames
    print(ws)
    # choose required sheet
    sheet = wb[sheetname]
    data = []
    # transfer the data to array
    for row in range(startrow, endrow):
        data.append(sheet.cell(row, col).value)
    return data

