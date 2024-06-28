# -*- coding: utf-8 -*-
"""
Created on Fri Dec 21 16:06:03 2018

@author: Balasubramaniam
"""

from openpyxl import load_workbook
filePath="F:/ANZ_Aug_2017/GDP.xlsx"
wb=load_workbook(filePath,read_only=False,data_only=True)

sheetNames=wb.sheetnames
print(sheetNames)

sheet=wb["GDP"]
data=[]
for row in range(7,50):
    data.append(sheet.cell(row=row,column=5).value)
    
    
import statistics
#average
print(statistics.mean(data))   
#sum
print(statistics.math.fsum(data))
#max
print(max(data))
#min
print(min(data))





